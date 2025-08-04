from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import re
from io import BytesIO
from utils.get_data import upload_file_to_bucket, get_next_facture_number, insert_facture, insert_ligne_facture, delete_facture_and_lines, download_template_from_storage
import convertapi
import streamlit as st
from datetime import datetime
import supabase  # Assumant que vous utilisez Supabase comme base de données


# Ces fonctions sont maintenant importées depuis utils.get_data


def invoice_maker(template_path=None,  # Paramètre conservé pour compatibilité mais non utilisé
                  output_path=None,       # Paramètre conservé pour compatibilité mais non utilisé
                  df=None,                # DataFrame ['produit','prix','quantité','tva']
                  client_info=None,       # [nom, adresse, contact]
                  date_prestation=None,   # ex. '2025-07-23'
                  total_ttc=None,
                  total_ht=None):
    """
    Crée une facture Excel/PDF et insère les données dans la base de données
    
    Returns:
        tuple: (filename, success) où success est True si toutes les insertions ont réussi
    """
    
    num_facture = get_next_facture_number()
    filename = f"facture_{num_facture}.xlsx"
    pdf_filename = filename.replace('.xlsx', '.pdf')
    
    try:
        # 1. Télécharger le template depuis Supabase Storage
        template_result = download_template_from_storage("template.xlsx")
        if not template_result["success"]:
            print(f"Erreur lors du téléchargement du template: {template_result['error']}")
            return pdf_filename, False
        
        # Créer un BytesIO depuis les données du template
        template_data = BytesIO(template_result["data"])
        
        # Charger le modèle et désactiver le mode template
        wb = load_workbook(template_data)
        wb.template = False
        ws = wb.active
        print(df)
        
        # 2. Injecter les infos client / date sur E9, E10, E11 et B16
        ws['E3']  = num_facture     # Numéro facture
        ws['E9']  = client_info["nom"]     # Nom client
        ws['E10'] = client_info["adresse_1"]     # Adresse client
        ws['E11'] = client_info["adresse_2"]     # Contact client
        if date_prestation != None : 
            ws['B16'] = str(date_prestation)    # Date de prestation
        else : 
            ws['A16'] = ""

        # 3. Récupérer le tableau "Tableau1"
        table = ws.tables["Tableau1"]
        start_ref, end_ref = table.ref.split(':')
        start_col, start_row = coordinate_from_string(start_ref)
        end_col,   end_row   = coordinate_from_string(end_ref)
        first_idx = column_index_from_string(start_col)
        last_idx  = column_index_from_string(end_col)

        # 4. Sauver formules et formats de la 1ère ligne de données
        model_row = start_row + 1
        formulas = {}
        formats  = {}
        for col in range(first_idx, last_idx + 1):
            c = ws.cell(row=model_row, column=col)
            if isinstance(c.value, str) and c.value.startswith('='):
                formulas[col] = c.value
            formats[col] = c.number_format

        # 5. Insérer len(df) lignes juste sous l'entête
        if len(df) > 1 : 
            ws.insert_rows(model_row+1, amount=len(df)-1)

        # 6. Remplir chaque nouvelle ligne
        for i, record in df.iterrows():
            r = model_row + i
            # Valeurs brutes dans l'ordre produit, prix, quantité, tva
            ws.cell(r, first_idx  , record["Produit"])  # produit
            ws.cell(r, first_idx+2, record["Prix unitaire"])  # prix
            ws.cell(r, first_idx+1, record["Quantité"])  # quantité
            ws.cell(r, first_idx+3, record["TVA"]/100)  # tva
            # Copier/adapter formulas
            for col, f in formulas.items():
                newf = re.sub(r'(\$?[A-Z]+\$?)(\d+)',
                              lambda m: f"{m.group(1)}{r}", f)
                ws.cell(r, col, newf)
            # Réappliquer formats
            for col, nf in formats.items():
                ws.cell(r, col).number_format = nf

        # 7. Mettre à jour la plage du tableau
        new_end = end_row + len(df)-1
        table.ref = f"{start_col}{start_row}:{end_col}{new_end}"
        
        # 8. Générer les fichiers Excel et PDF
        convertapi.api_credentials = st.secrets["CONVERTAPI"]

        file_content = BytesIO()
        wb.save(file_content)
        file_content.seek(0)

        # Upload XLSX
        result_xlsx = upload_file_to_bucket(file_content.getvalue(), filename)
        print(f"Upload XLSX: {result_xlsx}")

        # Sauvegarder xlsx temporairement pour conversion
        with open("data/temp.xlsx", 'wb') as f:
            f.write(file_content.getvalue())

        result = convertapi.convert('pdf', { 'File': 'data/temp.xlsx' })
        result.file.save('data/temp.pdf')
        
        # Upload PDF
        with open('data/temp.pdf', 'rb') as pdf_file:
            result_pdf = upload_file_to_bucket(pdf_file.read(), pdf_filename)
        print(f"Upload PDF: {result_pdf}")

        # 9. Insérer la facture principale dans la base de données
        print(f"Insertion facture - TTC: {total_ttc}, HT: {total_ht}")
        insertion_facture = insert_facture(
            num_facture=num_facture,
            date=datetime.now().strftime("%Y-%m-%d"),
            client_id=int(client_info["id"]),
            tot_ttc=float(total_ttc),
            tot_ht=float(total_ht),
            date_prestation=date_prestation if date_prestation is not None else None,
            paye=False
        )
        
        if not insertion_facture["success"]:
            print(f"Erreur lors de l'insertion de la facture principale: {insertion_facture['error']}")
            return pdf_filename, False

        # 10. Insérer toutes les lignes de facture
        lignes_inserees = []
        all_insertions_success = True
        
        for i, record in df.iterrows():
            ligne_result = insert_ligne_facture(
                num_facture=num_facture,
                produit=record["Produit"],
                quantite=record["Quantité"],
                prix_unitaire=record["Prix unitaire"],
                tva=record["TVA"]/100  # Convertir en décimal
            )
            print(ligne_result)
            
            lignes_inserees.append(ligne_result["success"])
            if not ligne_result["success"]:
                all_insertions_success = False
                print(f"Erreur lors de l'insertion de la ligne {i+1}: {record['Produit']} - {ligne_result['error']}")
                break
        
        # 11. Si une insertion a échoué, annuler toutes les insertions
        if not all_insertions_success:
            print("Une ou plusieurs insertions ont échoué. Annulation de toutes les insertions...")
            delete_result = delete_facture_and_lines(num_facture)
            if not delete_result["success"]:
                print(f"Erreur lors du rollback: {delete_result['error']}")
            return pdf_filename, False
        
        print(f"Toutes les insertions ont réussi pour la facture {num_facture}")
        return pdf_filename, True
        
    except Exception as e:
        print(f"Erreur générale dans invoice_maker: {e}")
        # En cas d'erreur, essayer de nettoyer
        try:
            delete_facture_and_lines(num_facture)
        except:
            pass
        return pdf_filename, False
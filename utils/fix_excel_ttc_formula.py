"""
Script pour corriger la formule du montant total TTC dans les fichiers Excel de factures.

Recherche la cellule "Montant Total TTC" et modifie la cellule 2 colonnes à droite
pour y mettre la formule: =SUMPRODUCT(Tableau1[Total HT])+SUMPRODUCT(Tableau1[Total HT],Tableau1[TVA])

Usage:
    python utils/fix_excel_ttc_formula.py [chemin_dossier_factures]
    
Exemple:
    python utils/fix_excel_ttc_formula.py ./factures_download
    python utils/fix_excel_ttc_formula.py ~/Documents/Factures_AG
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_cell_by_value(worksheet, search_text):
    """
    Recherche une cellule contenant un texte spécifique dans la feuille.
    
    Args:
        worksheet: Feuille de calcul openpyxl
        search_text: Texte à rechercher
    
    Returns:
        tuple: (row, column) de la cellule trouvée, ou (None, None) si non trouvée
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if search_text.lower() in cell.value.lower():
                    return cell.row, cell.column
    return None, None


def fix_ttc_formula_in_excel(file_path):
    """
    Corrige la formule TTC dans un fichier Excel de facture.
    
    Args:
        file_path: Chemin du fichier Excel à corriger
    
    Returns:
        dict: Résultat de l'opération
    """
    try:
        # Charger le workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Rechercher la cellule "Montant Total TTC"
        row, col = find_cell_by_value(ws, "Montant Total TTC")
        
        if row is None or col is None:
            return {
                "success": False,
                "error": "Cellule 'Montant Total TTC' non trouvée",
                "modified": False
            }
        
        # Calculer la position de la cellule cible (2 colonnes à droite)
        target_col = col + 2
        target_cell = ws.cell(row=row, column=target_col)
        
        # Formule à insérer (en anglais pour Excel)
        new_formula = "=SUMPRODUCT(Tableau1[Total HT])+SUMPRODUCT(Tableau1[Total HT],Tableau1[TVA])"
        
        # Vérifier si la formule est déjà correcte
        if target_cell.value == new_formula:
            return {
                "success": True,
                "error": None,
                "modified": False,
                "message": "Formule déjà correcte"
            }
        
        # Sauvegarder l'ancienne valeur pour log
        old_value = target_cell.value
        
        # Modifier la formule
        target_cell.value = new_formula
        
        # Sauvegarder le fichier
        wb.save(file_path)
        
        return {
            "success": True,
            "error": None,
            "modified": True,
            "old_value": old_value,
            "new_value": new_formula,
            "cell_position": f"{get_column_letter(target_col)}{row}"
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "modified": False
        }


def fix_all_excel_files(folder_path):
    """
    Corrige la formule TTC dans tous les fichiers Excel de factures du dossier.
    
    Args:
        folder_path: Chemin du dossier contenant les factures
    
    Returns:
        dict: Statistiques de l'opération
    """
    # Convertir en Path
    folder = Path(folder_path).expanduser().resolve()
    
    if not folder.exists():
        print(f"❌ Le dossier {folder} n'existe pas")
        return {
            "total": 0,
            "modified": 0,
            "skipped": 0,
            "failed": 0
        }
    
    print(f"\n📁 Dossier: {folder}")
    print(f"🔍 Recherche des fichiers Excel de factures...")
    
    # Trouver tous les fichiers Excel de factures
    excel_files = list(folder.glob("facture_*.xlsx"))
    
    if not excel_files:
        print("ℹ️  Aucun fichier Excel de facture trouvé.")
        return {
            "total": 0,
            "modified": 0,
            "skipped": 0,
            "failed": 0
        }
    
    print(f"\n📊 {len(excel_files)} fichier(s) Excel trouvé(s)")
    print(f"{'='*80}")
    
    # Statistiques
    stats = {
        "total": len(excel_files),
        "modified": 0,
        "skipped": 0,
        "failed": 0
    }
    
    # Traiter chaque fichier
    for excel_file in sorted(excel_files):
        filename = excel_file.name
        result = fix_ttc_formula_in_excel(excel_file)
        
        if result["success"]:
            if result["modified"]:
                stats["modified"] += 1
                print(f"✅ Modifié: {filename}")
                print(f"   📍 Cellule: {result['cell_position']}")
                print(f"   📝 Ancienne valeur: {result.get('old_value', 'N/A')}")
                print(f"   ✨ Nouvelle formule: {result['new_value']}")
            else:
                stats["skipped"] += 1
                print(f"⏭️  Déjà correct: {filename}")
        else:
            stats["failed"] += 1
            print(f"❌ Erreur sur {filename}: {result['error']}")
    
    # Afficher le résumé
    print(f"\n{'='*80}")
    print(f"📊 RÉSUMÉ")
    print(f"{'='*80}")
    print(f"  Total de fichiers:      {stats['total']}")
    print(f"  ✅ Modifiés:            {stats['modified']}")
    print(f"  ⏭️  Déjà corrects:       {stats['skipped']}")
    print(f"  ❌ Erreurs:             {stats['failed']}")
    print(f"{'='*80}\n")
    
    return stats


def main():
    """
    Point d'entrée principal du script.
    """
    # Vérifier les arguments
    if len(sys.argv) > 1:
        folder_path = sys.argv[1]
    else:
        # Dossier par défaut
        folder_path = "/Users/sajidasarumugadas/Library/CloudStorage/OneDrive-Personnel/sarl ASD/Compta/2025/Factures_emises"
        print(f"ℹ️  Aucun dossier spécifié, utilisation du dossier par défaut: {folder_path}")
        print(f"💡 Usage: python utils/fix_excel_ttc_formula.py [chemin_dossier]\n")
    
    try:
        stats = fix_all_excel_files(folder_path)
        
        # Code de sortie selon les résultats
        if stats["failed"] > 0:
            sys.exit(1)
        else:
            sys.exit(0)
            
    except KeyboardInterrupt:
        print("\n\n⚠️  Traitement interrompu par l'utilisateur")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Erreur fatale: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
